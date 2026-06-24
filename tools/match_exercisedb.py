#!/usr/bin/env python3
"""Match FitForge Excel catalog to ExerciseDB (oss.exercisedb.dev)."""

from __future__ import annotations

import csv
import json
import re
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
from rapidfuzz import fuzz

ROOT = Path(__file__).resolve().parents[1]
EXCEL = Path(r"c:\Users\xemil\Downloads\fitforge_exercise_catalog.xlsx")
OUT_DIR = ROOT / "tools" / "exercise_match_output"
API = "https://oss.exercisedb.dev/api/v1/exercises"

EQUIPMENT_MAP = {
    "barbell": "barbell",
    "dumbbell": "dumbbell",
    "cable": "cable",
    "machine": "leverage machine",
    "bodyweight": "body weight",
    "smith machine": "smith machine",
}

GROUP_BODY_PARTS = {
    "chest": {"chest"},
    "back": {"back", "lats"},
    "legs": {"upper legs", "lower legs", "quads", "hamstrings", "calves"},
    "shoulders": {"shoulders"},
    "biceps": {"upper arms"},
    "triceps": {"upper arms"},
    "glutes": {"upper legs", "glutes"},
    "calves": {"lower legs", "calves"},
}

MUSCLE_HINTS = {
    "chest": {"chest", "pectorals", "pecs", "upper chest", "lower chest"},
    "lats": {"lats", "latissimus", "back"},
    "triceps": {"triceps"},
    "biceps": {"biceps"},
    "quads": {"quads", "quadriceps"},
    "glutes": {"glutes", "gluteus"},
    "calves": {"calves", "gastrocnemius", "soleus"},
    "front delts": {"delts", "deltoids", "shoulders", "anterior deltoid"},
    "rear delts": {"delts", "deltoids", "shoulders", "rear deltoid"},
    "side delts": {"delts", "deltoids", "shoulders", "lateral deltoid"},
    "upper chest": {"upper chest", "pectorals", "chest"},
    "lower chest": {"lower chest", "pectorals", "chest"},
    "hamstrings": {"hamstrings"},
    "abs": {"abs", "abdominals", "core"},
    "forearms": {"forearms"},
    "traps": {"traps", "trapezius"},
    "rhomboids": {"rhomboids", "upper back"},
    "erector spinae": {"spine", "lower back", "erector"},
}


def slug(text: str) -> str:
    text = text.lower().strip()
    text = (
        text.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def make_ff_id(group: str, name_en: str) -> str:
    return "ff_" + slug(f"{group} {name_en}").replace(" ", "_")


def load_excel() -> list[dict]:
    wb = load_workbook(EXCEL, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    items = []
    for row in rows[1:]:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        group = str(row[0] or "").strip()
        name_en = str(row[1] or "").strip()
        items.append(
            {
                "ff_id": make_ff_id(group, name_en),
                "group": group,
                "name_en": name_en,
                "name_es": str(row[2] or "").strip(),
                "equipment": str(row[3] or "").strip(),
                "load_mode": str(row[4] or "").strip(),
                "independent_load": bool(row[5]),
                "unilateral": bool(row[6]),
                "primary_muscle": str(row[7] or "").strip(),
                "secondary_muscles": str(row[8] or "").strip(),
                "notes": str(row[9] or "").strip() if len(row) > 9 and row[9] else "",
            }
        )
    return items


MIRROR_URL = (
    "https://raw.githubusercontent.com/hasaneyldrm/exercises-dataset/main/data/exercises.json"
)


def _extract_exercise_id(gif_url: str | None) -> str | None:
    if not gif_url:
        return None
    match = re.search(r"([A-Za-z0-9]{6,12})\.gif$", gif_url.replace("\\", "/"))
    return match.group(1) if match else None


def _normalize_mirror_item(item: dict) -> dict | None:
    exercise_id = _extract_exercise_id(item.get("gif_url"))
    name = (item.get("name") or "").strip()
    if not exercise_id or not name:
        return None

    equipment = item.get("equipment") or ""
    body_part = item.get("body_part") or item.get("category") or ""
    target = item.get("target") or item.get("muscle_group") or ""
    secondary = item.get("secondary_muscles") or []

    if isinstance(secondary, str):
        secondary = [s.strip() for s in secondary.split(",") if s.strip()]

    return {
        "exerciseId": exercise_id,
        "name": name,
        "gifUrl": f"https://static.exercisedb.dev/media/{exercise_id}.gif",
        "bodyParts": [body_part] if body_part else [],
        "equipments": [equipment] if equipment else [],
        "targetMuscles": [target] if target else [],
        "secondaryMuscles": secondary if isinstance(secondary, list) else [],
        "source": "mirror",
    }


def fetch_exercisedb_mirror() -> list[dict]:
    print("Downloading ExerciseDB mirror dataset...", flush=True)
    response = requests.get(MIRROR_URL, timeout=120)
    response.raise_for_status()
    raw = response.json()
    normalized: list[dict] = []
    for item in raw:
        parsed = _normalize_mirror_item(item)
        if parsed is not None:
            normalized.append(parsed)
    print(f"Mirror normalized: {len(normalized)} exercises", flush=True)
    return normalized


def fetch_exercisedb_api() -> list[dict]:
    by_id: dict[str, dict] = {}
    cursor: str | None = None
    seen_cursors: set[str] = set()
    page = 0

    while page < 80:
        params: dict[str, str | int] = {"limit": 25}
        if cursor:
            if cursor in seen_cursors:
                print("  cursor repeated, stopping", flush=True)
                break
            seen_cursors.add(cursor)
            params["cursor"] = cursor

        body = _get_page(params)
        batch = body["data"]
        meta = body["meta"]
        new_count = 0
        for item in batch:
            if item["exerciseId"] not in by_id:
                new_count += 1
            by_id[item["exerciseId"]] = item

        page += 1
        print(
            f"  page {page}: batch={len(batch)} new={new_count} total={len(by_id)} "
            f"target={meta.get('total', '?')}",
            flush=True,
        )

        if not batch or new_count == 0:
            break
        if not meta.get("hasNextPage"):
            break
        cursor = meta.get("nextCursor")
        time.sleep(2.5)

    return list(by_id.values())


def _get_page(params: dict[str, str | int]) -> dict:
    for attempt in range(8):
        response = requests.get(API, params=params, timeout=60)
        if response.status_code == 429:
            wait = min(30, 2 ** attempt)
            print(f"  rate limited, waiting {wait}s...", flush=True)
            time.sleep(wait)
            continue
        response.raise_for_status()
        return response.json()
    raise RuntimeError(f"ExerciseDB rate limit for params={params}")


def normalize_equipment(value: str) -> str:
    return EQUIPMENT_MAP.get(slug(value), slug(value))


def equipment_score(excel_eq: str, edb_equipments: list[str]) -> float:
    target = normalize_equipment(excel_eq)
    edb = [slug(e) for e in edb_equipments]
    if not edb:
        return 0.0
    if target in edb:
        return 1.0
    if target == "leverage machine" and any("machine" in e for e in edb):
        return 0.85
    if target == "body weight" and "body weight" in edb:
        return 1.0
    return max(fuzz.partial_ratio(target, e) for e in edb) / 100.0


def body_part_score(group: str, body_parts: list[str]) -> float:
    expected = GROUP_BODY_PARTS.get(slug(group), set())
    if not expected or not body_parts:
        return 0.5
    edb = {slug(p) for p in body_parts}
    if expected & edb:
        return 1.0
    return 0.35


def muscle_score(primary: str, target_muscles: list[str], secondary_muscles: list[str]) -> float:
    hints = MUSCLE_HINTS.get(slug(primary), {slug(primary)})
    pool = [slug(m) for m in [*target_muscles, *secondary_muscles]]
    if not pool:
        return 0.5
    best = 0.0
    for hint in hints:
        for muscle in pool:
            if hint in muscle or muscle in hint:
                best = max(best, 1.0)
            else:
                best = max(best, fuzz.partial_ratio(hint, muscle) / 100.0)
    return best


VARIANT_TERMS = [
    "incline",
    "decline",
    "seated",
    "standing",
    "single arm",
    "one arm",
    "alternating",
    "reverse",
    "wide grip",
    "close grip",
    "neutral grip",
    "supinated",
    "pronated",
    "smith",
    "machine",
    "cable",
    "barbell",
    "dumbbell",
    "bodyweight",
    "walking",
    "romanian",
    "conventional",
    "front",
    "back",
    "lateral",
    "rear",
    "high to low",
    "low to high",
]


DISTINCTIVE_TERMS = [
    "archer",
    "diamond",
    "single leg",
    "one leg",
    "monster",
    "frog",
    "fire hydrant",
    "wall walk",
    "face pull",
    "meadows",
]


def variant_penalty(excel_name: str, edb_name: str) -> float:
    a = slug(excel_name)
    b = slug(edb_name)
    penalty = 0.0
    for term in VARIANT_TERMS:
        in_a = term in a
        in_b = term in b
        if in_a != in_b:
            penalty += 0.12 if term in {"incline", "decline", "seated", "standing", "romanian", "conventional", "front", "rear", "lateral"} else 0.08
    for term in DISTINCTIVE_TERMS:
        in_a = term in a
        in_b = term in b
        if in_b and not in_a:
            penalty += 0.35
        if in_a and not in_b:
            penalty += 0.25
    return min(penalty, 0.65)


def name_score(excel_name: str, edb_name: str) -> float:
    a = slug(excel_name)
    b = slug(edb_name)
    base = max(
        fuzz.token_sort_ratio(a, b),
        fuzz.token_set_ratio(a, b),
        fuzz.partial_ratio(a, b),
    ) / 100.0
    return max(0.0, base - variant_penalty(excel_name, edb_name))


def score_pair(excel: dict, edb: dict) -> tuple[float, dict[str, float]]:
    n = name_score(excel["name_en"], edb["name"])
    e = equipment_score(excel["equipment"], edb.get("equipments") or [])
    b = body_part_score(excel["group"], edb.get("bodyParts") or [])
    m = muscle_score(
        excel["primary_muscle"],
        edb.get("targetMuscles") or [],
        edb.get("secondaryMuscles") or [],
    )
    total = 0.55 * n + 0.25 * e + 0.10 * b + 0.10 * m
    return total, {"name": n, "equipment": e, "body_part": b, "muscle": m}


def confidence_label(score: float) -> str:
    if score >= 0.82:
        return "high"
    if score >= 0.68:
        return "medium"
    if score >= 0.52:
        return "low"
    return "none"


def match_catalog(excel_items: list[dict], edb_items: list[dict]) -> dict:
    matches = []
    unmatched = []

    for excel in excel_items:
        scored = []
        for edb in edb_items:
            total, breakdown = score_pair(excel, edb)
            scored.append((total, breakdown, edb))
        scored.sort(key=lambda x: x[0], reverse=True)

        best_score, best_breakdown, best = scored[0]
        second_score = scored[1][0] if len(scored) > 1 else 0.0
        margin = best_score - second_score
        label = confidence_label(best_score)

        entry = {
            "ff_id": excel["ff_id"],
            "name_en": excel["name_en"],
            "name_es": excel["name_es"],
            "group": excel["group"],
            "equipment": excel["equipment"],
            "primary_muscle": excel["primary_muscle"],
            "match_confidence": label,
            "match_score": round(best_score, 4),
            "score_margin": round(margin, 4),
            "score_breakdown": {k: round(v, 4) for k, v in best_breakdown.items()},
            "exercisedb": {
                "exerciseId": best["exerciseId"],
                "name": best["name"],
                "gifUrl": best["gifUrl"],
                "equipments": best.get("equipments"),
                "bodyParts": best.get("bodyParts"),
                "targetMuscles": best.get("targetMuscles"),
            },
            "alternatives": [
                {
                    "exerciseId": alt["exerciseId"],
                    "name": alt["name"],
                    "gifUrl": alt["gifUrl"],
                    "match_score": round(score, 4),
                }
                for score, _, alt in scored[1:4]
            ],
        }

        if label == "none":
            unmatched.append(entry)
        else:
            matches.append(entry)

    summary = {
        "excel_total": len(excel_items),
        "exercisedb_total": len(edb_items),
        "matched_high": sum(1 for m in matches if m["match_confidence"] == "high"),
        "matched_medium": sum(1 for m in matches if m["match_confidence"] == "medium"),
        "matched_low": sum(1 for m in matches if m["match_confidence"] == "low"),
        "unmatched": len(unmatched),
    }
    return {"summary": summary, "matches": matches, "unmatched": unmatched}


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    excel_items = load_excel()
    print(f"Excel: {len(excel_items)} exercises")

    raw_path = OUT_DIR / "exercisedb_raw.json"
    if raw_path.exists():
        edb_items = json.loads(raw_path.read_text(encoding="utf-8"))
        print(f"ExerciseDB: {len(edb_items)} exercises (cached)", flush=True)
    else:
        api_items = fetch_exercisedb_api()
        if len(api_items) >= 200:
            edb_items = api_items
            print(f"ExerciseDB: {len(edb_items)} exercises (API)", flush=True)
        else:
            print(
                f"API returned only {len(api_items)} exercises; using public mirror fallback.",
                flush=True,
            )
            edb_items = fetch_exercisedb_mirror()
        raw_path.write_text(json.dumps(edb_items, ensure_ascii=False, indent=2), encoding="utf-8")

    result = match_catalog(excel_items, edb_items)
    full = {"generated_at": time.strftime("%Y-%m-%dT%H:%M:%S"), **result}
    (OUT_DIR / "exercise_match_results.json").write_text(
        json.dumps(full, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # Compact mapping for app integration review (high + medium only)
    mapping = {
        m["ff_id"]: {
            "exerciseId": m["exercisedb"]["exerciseId"],
            "gifUrl": m["exercisedb"]["gifUrl"],
            "confidence": m["match_confidence"],
            "score": m["match_score"],
            "edb_name": m["exercisedb"]["name"],
        }
        for m in result["matches"]
        if m["match_confidence"] in {"high", "medium"}
    }
    (OUT_DIR / "exercise_match_mapping.json").write_text(
        json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    all_rows = sorted(
        [*result["matches"], *result["unmatched"]],
        key=lambda m: m["match_score"],
        reverse=True,
    )
    csv_path = OUT_DIR / "exercise_match_review.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "ff_id",
                "name_en",
                "name_es",
                "equipment",
                "confidence",
                "score",
                "edb_exerciseId",
                "edb_name",
                "gifUrl",
                "alt1_name",
                "alt1_score",
            ],
        )
        writer.writeheader()
        for m in all_rows:
            alt = m["alternatives"][0] if m["alternatives"] else {}
            writer.writerow(
                {
                    "ff_id": m["ff_id"],
                    "name_en": m["name_en"],
                    "name_es": m["name_es"],
                    "equipment": m["equipment"],
                    "confidence": m["match_confidence"],
                    "score": m["match_score"],
                    "edb_exerciseId": m["exercisedb"]["exerciseId"],
                    "edb_name": m["exercisedb"]["name"],
                    "gifUrl": m["exercisedb"]["gifUrl"],
                    "alt1_name": alt.get("name", ""),
                    "alt1_score": alt.get("match_score", ""),
                }
            )

    s = result["summary"]
    print("\n=== MATCH SUMMARY ===")
    print(json.dumps(s, indent=2))
    print(f"\nOutput: {OUT_DIR}")


if __name__ == "__main__":
    main()
