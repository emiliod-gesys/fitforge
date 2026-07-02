#!/usr/bin/env python3
"""Build FitForge bundled exercise catalog from Excel + ExerciseDB mapping."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL = Path(r"c:\Users\xemil\Downloads\fitforge_exercise_catalog.xlsx")
MATCH_RESULTS = ROOT / "tools" / "exercise_match_output" / "exercise_match_results.json"
OUT = ROOT / "assets" / "data" / "exercise_catalog.json"
LEGACY_OUT = ROOT / "assets" / "data" / "legacy" / "exercise_catalog_seed_v1.json"

GROUP_ES = {
    "Chest": "Pecho",
    "Back": "Espalda",
    "Legs": "Piernas",
    "Shoulders": "Hombros",
    "Biceps": "Bíceps",
    "Triceps": "Tríceps",
    "Glutes": "Glúteos",
    "Cardio": "Cardio",
    "Calves": "Pantorrillas",
    "Abs": "Abdominales",
}

GROUP_EN = {
    "Chest": "Chest",
    "Back": "Back",
    "Legs": "Legs",
    "Shoulders": "Shoulders",
    "Biceps": "Biceps",
    "Triceps": "Triceps",
    "Glutes": "Glutes",
    "Calves": "Calves",
    "Cardio": "Cardio",
    "Abs": "Abs",
}

EQUIPMENT_ES = {
    "Barbell": "Barra",
    "Dumbbell": "Mancuernas",
    "Cable": "Polea",
    "Machine": "Máquina",
    "Bodyweight": "Peso corporal",
    "Smith Machine": "Smith",
    "Cardio Machine": "Máquina cardio",
    "Outdoor": "Aire libre",
}

EQUIPMENT_EN = {
    "Barbell": "Barbell",
    "Dumbbell": "Dumbbell",
    "Cable": "Cable",
    "Machine": "Machine",
    "Bodyweight": "Bodyweight",
    "Smith Machine": "Smith Machine",
    "Cardio Machine": "Cardio machine",
    "Outdoor": "Outdoor",
}

MUSCLE_ES = {
    "Chest": "Pecho",
    "Upper Chest": "Pecho superior",
    "Lower Chest": "Pecho inferior",
    "Lats": "Dorsales",
    "Upper Back": "Espalda alta",
    "Back": "Espalda",
    "Lower Traps": "Trapecio inferior",
    "Rhomboids": "Romboides",
    "Front Delts": "Deltoides anterior",
    "Side Delts": "Deltoides lateral",
    "Rear Delts": "Deltoides posterior",
    "Shoulders": "Hombros",
    "Biceps": "Bíceps",
    "Brachialis": "Braquial",
    "Triceps": "Tríceps",
    "Quads": "Cuádriceps",
    "Hamstrings": "Isquios",
    "Adductors": "Aductores",
    "Glutes": "Glúteos",
    "Glute Medius": "Glúteo medio",
    "Calves": "Pantorrillas",
    "Forearms": "Antebrazos",
    "Abs": "Abdominales",
    "Core": "Core",
    "Traps": "Trapecios",
    "Arms": "Brazos",
    "Cardio": "Cardio",
}

MUSCLE_EN = {
    "Chest": "Chest",
    "Upper Chest": "Upper chest",
    "Lower Chest": "Lower chest",
    "Lats": "Lats",
    "Upper Back": "Upper back",
    "Back": "Back",
    "Lower Traps": "Lower traps",
    "Rhomboids": "Rhomboids",
    "Front Delts": "Front delts",
    "Side Delts": "Side delts",
    "Rear Delts": "Rear delts",
    "Shoulders": "Shoulders",
    "Biceps": "Biceps",
    "Brachialis": "Brachialis",
    "Triceps": "Triceps",
    "Quads": "Quads",
    "Hamstrings": "Hamstrings",
    "Adductors": "Adductors",
    "Glutes": "Glutes",
    "Glute Medius": "Glute medius",
    "Calves": "Calves",
    "Forearms": "Forearms",
    "Abs": "Abs",
    "Core": "Core",
    "Traps": "Traps",
    "Cardio": "Cardio",
}

# Reemplaza saltar cuerda bajo Pantorrillas por la entrada en Cardio.
EXCLUDE_FF_IDS = {"ff_calves_jump_rope"}

# ExerciseDB no tiene hip thrust con barra/mancuerna; el fuzzy match eligió lunges/squats.
# Usamos el movimiento más cercano (puente/hip thrust con espalda en banco).
EXERCISEDB_OVERRIDES: dict[str, str] = {
    "ff_glutes_dumbbell_hip_thrust": "aWedzZX",  # glute bridge two legs on bench
    "ff_shoulders_reverse_pec_deck_machine": "myfUsKf",  # lever seated reverse fly
}

# Ilustraciones propias cuando ExerciseDB no tiene un GIF fiel.
LOCAL_IMAGE_OVERRIDES: dict[str, str] = {
    "ff_legs_hack_squat_machine": "assets/exercises/ff_legs_hack_squat_machine.png",
    "ff_glutes_barbell_hip_thrust": "assets/exercises/ff_glutes_barbell_hip_thrust.webp",
    "ff_glutes_hip_thrust_machine": "assets/exercises/ff_glutes_hip_thrust_machine.jpg",
    "ff_glutes_smith_machine_hip_thrust": "assets/exercises/ff_glutes_smith_machine_hip_thrust.jpg",
    "ff_back_chest_supported_row_machine": "assets/exercises/ff_back_chest_supported_row_machine.png",
    "ff_chest_incline_chest_press_machine": "assets/exercises/ff_chest_incline_chest_press_machine.png",
    "ff_triceps_machine_dip": "assets/exercises/ff_triceps_machine_dip.png",
    "ff_chest_pec_deck_fly_machine": "assets/exercises/ff_chest_pec_deck_fly_machine.png",
    "ff_cardio_rowing_machine": "assets/exercises/ff_cardio_rowing_machine.png",
    "ff_back_single_arm_dumbbell_row": "assets/exercises/ff_back_single_arm_dumbbell_row.png",
    "ff_cardio_air_bike": "assets/exercises/ff_cardio_air_bike.png",
    "ff_shoulders_face_pull": "assets/exercises/ff_shoulders_face_pull.png",
    "ff_shoulders_shoulder_press_machine": "assets/exercises/ff_shoulders_shoulder_press_machine.png",
    "ff_biceps_single_arm_cable_curl": "assets/exercises/ff_biceps_single_arm_cable_curl.png",
    "ff_biceps_dumbbell_reverse_spider_curl": "assets/exercises/ff_biceps_dumbbell_reverse_spider_curl.png",
    "ff_calves_smith_machine_calf_raise": "assets/exercises/ff_calves_smith_machine_calf_raise.png",
    "ff_legs_standing_leg_curl_machine": "assets/exercises/ff_legs_standing_leg_curl_machine.png",
    "ff_legs_seated_leg_curl_machine": "assets/exercises/ff_legs_seated_leg_curl_machine.png",
    "ff_legs_belt_squat_machine": "assets/exercises/ff_legs_belt_squat_machine.png",
    "ff_glutes_cable_kickback": "assets/exercises/ff_glutes_cable_kickback.png",
    "ff_glutes_glute_drive_machine": "assets/exercises/ff_glutes_glute_drive_machine.png",
    "ff_legs_adductor_machine": "assets/exercises/ff_legs_adductor_machine.png",
    "ff_calves_leg_press_calf_raise_machine": "assets/exercises/ff_calves_leg_press_calf_raise_machine.png",
    "ff_calves_seated_calf_raise_machine": "assets/exercises/ff_calves_seated_calf_raise_machine.png",
    "ff_glutes_reverse_hyper_machine": "assets/exercises/ff_glutes_reverse_hyper_machine.png",
}

# Ejercicios añadidos fuera del Excel (evita duplicados por id al regenerar).
CATALOG_SUPPLEMENT = [
    {
        "group": "Biceps",
        "name_en": "Dumbbell Reverse Spider Curl",
        "name_es": "Curl araña inverso con mancuernas",
        "equipment": "Dumbbell",
        "load_mode": "dual_load",
        "per_arm_weight": True,
        "unilateral": False,
        "primary": "Biceps",
        "secondary": ["Forearms"],
        "exercisedb_id": "6sMAmNv",
    },
    {
        "group": "Legs",
        "name_en": "Adductor Machine",
        "name_es": "Aducción de cadera en máquina",
        "equipment": "Machine",
        "load_mode": "machine_stack",
        "per_arm_weight": False,
        "unilateral": False,
        "primary": "Adductors",
        "secondary": ["Glutes"],
        "exercisedb_id": "oHsrypV",
    },
    # Abdominales (no estaban en el Excel; GIFs de ExerciseDB)
    {
        "group": "Abs",
        "name_en": "Crunch",
        "name_es": "Crunch abdominal",
        "equipment": "Bodyweight",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "TFqbd8t",
    },
    {
        "group": "Abs",
        "name_en": "Plank",
        "name_es": "Plancha abdominal",
        "equipment": "Bodyweight",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "VBAWRPG",
    },
    {
        "group": "Abs",
        "name_en": "Hanging Leg Raise",
        "name_es": "Elevación de piernas colgado",
        "equipment": "Bodyweight",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "I3tsCnC",
    },
    {
        "group": "Abs",
        "name_en": "Cable Crunch",
        "name_es": "Crunch en polea",
        "equipment": "Cable",
        "load_mode": "single_load",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "WW95auq",
    },
    {
        "group": "Abs",
        "name_en": "Bicycle Crunch",
        "name_es": "Crunch bicicleta",
        "equipment": "Bodyweight",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "tZkGYZ9",
    },
    {
        "group": "Abs",
        "name_en": "Russian Twist",
        "name_es": "Giros rusos",
        "equipment": "Bodyweight",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "XVDdcoj",
    },
    {
        "group": "Abs",
        "name_en": "Dead Bug",
        "name_es": "Dead bug",
        "equipment": "Bodyweight",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "iny3m5y",
    },
    {
        "group": "Abs",
        "name_en": "Ab Wheel Rollout",
        "name_es": "Rueda abdominal",
        "equipment": "Bodyweight",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "NAgVB3t",
    },
    {
        "group": "Abs",
        "name_en": "Lying Leg Raise",
        "name_es": "Elevación de piernas acostado",
        "equipment": "Bodyweight",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "WhuFnR7",
    },
    {
        "group": "Abs",
        "name_en": "Sit-Up",
        "name_es": "Abdominal completo",
        "equipment": "Bodyweight",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "QLL2gdc",
    },
    # Abdominales en máquina y polea
    {
        "group": "Abs",
        "name_en": "Ab Crunch Machine",
        "name_es": "Crunch abdominal en máquina",
        "equipment": "Machine",
        "load_mode": "machine_stack",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "ZnJHhMk",
    },
    {
        "group": "Abs",
        "name_en": "Captain's Chair Leg Raise",
        "name_es": "Elevación de piernas en silla romana",
        "equipment": "Machine",
        "load_mode": "bodyweight",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "weoDEpH",
    },
    {
        "group": "Abs",
        "name_en": "Machine Seated Leg Raise Crunch",
        "name_es": "Crunch con elevación de piernas en máquina",
        "equipment": "Machine",
        "load_mode": "machine_stack",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "PQ2AtC3",
    },
    {
        "group": "Abs",
        "name_en": "Cable Standing Crunch",
        "name_es": "Crunch de pie en polea",
        "equipment": "Cable",
        "load_mode": "single_load",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "jpgqxiS",
    },
    {
        "group": "Abs",
        "name_en": "Cable Reverse Crunch",
        "name_es": "Crunch inverso en polea",
        "equipment": "Cable",
        "load_mode": "single_load",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "RqOtqD7",
    },
    {
        "group": "Abs",
        "name_en": "Cable Side Bend",
        "name_es": "Flexión lateral en polea",
        "equipment": "Cable",
        "load_mode": "single_load",
        "unilateral": True,
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "wPypxFY",
    },
    {
        "group": "Abs",
        "name_en": "Cable Wood Chop",
        "name_es": "Leñador en polea",
        "equipment": "Cable",
        "load_mode": "single_load",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "fhZQPlV",
    },
    {
        "group": "Abs",
        "name_en": "Cable Seated Twist",
        "name_es": "Giro sentado en polea",
        "equipment": "Cable",
        "load_mode": "single_load",
        "primary": "Abs",
        "secondary": [],
        "exercisedb_id": "UEjSrKI",
    },
]

CARDIO_SUPPLEMENT = [
    # Máquinas (hoja Cardio del Excel)
    {
        "name_en": "Treadmill",
        "name_es": "Caminadora",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Legs", "Glutes"],
        "cardio_preset": "treadmill",
        "exercisedb_id": "rjiM4L3",
    },
    {
        "name_en": "Elliptical Trainer",
        "name_es": "Elíptica",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Quads", "Glutes", "Calves"],
        "cardio_preset": "elliptical",
        "exercisedb_id": "rjtuP6X",
    },
    {
        "name_en": "Stationary Bike",
        "name_es": "Bicicleta estática",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Quads", "Glutes", "Calves"],
        "cardio_preset": "bike",
        "exercisedb_id": "H1PESYI",
    },
    {
        "name_en": "Recumbent Bike",
        "name_es": "Bicicleta reclinada",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Quads", "Glutes", "Calves"],
        "cardio_preset": "bike",
        "exercisedb_id": "a8VDgLw",
    },
    {
        "name_en": "Stair Climber",
        "name_es": "Escaladora",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Glutes", "Quads", "Calves"],
        "cardio_preset": "stairClimber",
        "exercisedb_id": "j9Q5crt",
    },
    {
        "name_en": "StepMill",
        "name_es": "Escaleras infinitas",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Glutes", "Quads", "Calves"],
        "cardio_preset": "stairClimber",
        "exercisedb_id": "j9Q5crt",
    },
    {
        "name_en": "Rowing Machine",
        "name_es": "Remo ergómetro",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Back", "Legs", "Biceps"],
        "cardio_preset": "rowing",
        "exercisedb_id": None,
    },
    {
        "name_en": "Ski Erg",
        "name_es": "Ski ergómetro",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Lats", "Triceps", "Core"],
        "cardio_preset": "custom",
        "exercisedb_id": "vpQaQkH",
    },
    {
        "name_en": "Air Bike",
        "name_es": "Bicicleta de aire",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Legs", "Shoulders", "Arms"],
        "cardio_preset": "bike",
        "exercisedb_id": "1ZFqTDN",
    },
    {
        "name_en": "Arc Trainer",
        "name_es": "Arc Trainer",
        "equipment": "Cardio Machine",
        "load_mode": "cardio_machine",
        "secondary": ["Glutes", "Quads", "Calves"],
        "cardio_preset": "elliptical",
        "exercisedb_id": "XSCHmiI",
    },
    # Cardio tradicional al aire libre
    {
        "name_en": "Outdoor Walking",
        "name_es": "Caminar al aire libre",
        "equipment": "Outdoor",
        "load_mode": "cardio_outdoor",
        "secondary": ["Legs", "Glutes"],
        "cardio_preset": "treadmill",
        "exercisedb_id": "CcWEoWV",
        "aliases_es": ["Caminata", "Caminar"],
        "aliases_en": ["Walking", "Brisk walk"],
    },
    {
        "name_en": "Outdoor Running",
        "name_es": "Correr al aire libre",
        "equipment": "Outdoor",
        "load_mode": "cardio_outdoor",
        "secondary": ["Legs", "Glutes", "Calves"],
        "cardio_preset": "treadmill",
        "exercisedb_id": "oLrKqDH",
        "aliases_es": ["Correr", "Trote", "Jogging"],
        "aliases_en": ["Running", "Jogging"],
    },
    {
        "name_en": "Outdoor Cycling",
        "name_es": "Ciclismo al aire libre",
        "equipment": "Outdoor",
        "load_mode": "cardio_outdoor",
        "secondary": ["Quads", "Glutes", "Calves"],
        "cardio_preset": "bike",
        "exercisedb_id": "km2Ljzj",
        "aliases_es": ["Bicicleta", "Bici", "Ciclismo"],
        "aliases_en": ["Cycling", "Road bike"],
    },
    {
        "name_en": "Jump Rope",
        "name_es": "Saltar cuerda",
        "equipment": "Outdoor",
        "load_mode": "cardio_outdoor",
        "secondary": ["Calves", "Legs"],
        "cardio_preset": "custom",
        "exercisedb_id": "e1e76I2",
        "aliases_es": ["Comba", "Cuerda"],
    },
    {
        "name_en": "Hiking",
        "name_es": "Senderismo",
        "equipment": "Outdoor",
        "load_mode": "cardio_outdoor",
        "secondary": ["Legs", "Glutes", "Calves"],
        "cardio_preset": "treadmill",
        "exercisedb_id": "rjiM4L3",
        "aliases_es": ["Trekking", "Montaña"],
        "aliases_en": ["Trekking"],
    },
    {
        "name_en": "Swimming",
        "name_es": "Natación",
        "equipment": "Outdoor",
        "load_mode": "cardio_outdoor",
        "secondary": ["Back", "Shoulders", "Core"],
        "cardio_preset": "custom",
        "exercisedb_id": "SP3hUez",
        "aliases_es": ["Nadar", "Piscina"],
        "aliases_en": ["Swim"],
    },
]


def slug(text: str) -> str:
    text = text.lower().strip()
    text = (
        text.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def make_ff_id(group: str, name_en: str) -> str:
    return "ff_" + slug(f"{group} {name_en}").replace(" ", "_")


def split_secondary(value: str | None) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def localize_muscle(name: str, lang: str) -> str:
    if lang == "es":
        return MUSCLE_ES.get(name, name)
    return MUSCLE_EN.get(name, name)


def load_matches() -> dict[str, dict]:
    data = json.loads(MATCH_RESULTS.read_text(encoding="utf-8"))
    mapping: dict[str, dict] = {}
    for key in ("matches", "unmatched"):
        for item in data.get(key, []):
            mapping[item["ff_id"]] = item
    return mapping


def infer_logging(
    ff_id: str, group: str, name_en: str, load_mode: str, cardio_preset: str | None = None
) -> tuple[str, str | None]:
    if group == "Cardio" or load_mode in {"cardio_machine", "cardio_outdoor"}:
        return "cardio", cardio_preset or "custom"
    return "strength", None


def gif_url(exercisedb_id: str | None) -> str | None:
    if not exercisedb_id:
        return None
    return f"https://static.exercisedb.dev/media/{exercisedb_id}.gif"


def build_strength_supplement_entry(spec: dict) -> dict:
    group = spec["group"]
    name_en = spec["name_en"]
    ff_id = make_ff_id(group, name_en)
    load_mode = spec["load_mode"]
    secondary = spec.get("secondary", [])
    primary = spec["primary"]
    logging_type, cardio_preset = infer_logging(ff_id, group, name_en, load_mode)

    per_arm = spec.get("per_arm_weight", load_mode == "dual_load")
    unilateral = spec.get("unilateral", False)
    weight_optional = load_mode in {"bodyweight", "assisted_bodyweight"}

    exercisedb_id = spec.get("exercisedb_id")
    image_url = gif_url(exercisedb_id)
    match_confidence = "curated"
    match_score = None

    if ff_id in LOCAL_IMAGE_OVERRIDES:
        image_url = LOCAL_IMAGE_OVERRIDES[ff_id]
        exercisedb_id = None
    elif ff_id in EXERCISEDB_OVERRIDES:
        exercisedb_id = EXERCISEDB_OVERRIDES[ff_id]
        image_url = gif_url(exercisedb_id)

    primary_es = localize_muscle(primary, "es")
    primary_en = localize_muscle(primary, "en")
    secondary_es = [localize_muscle(m, "es") for m in secondary]
    secondary_en = [localize_muscle(m, "en") for m in secondary]

    return {
        "id": ff_id,
        "loggingType": logging_type,
        "loadMode": load_mode,
        "perArmWeight": per_arm,
        "unilateral": unilateral,
        "weightOptional": weight_optional,
        "cardioPreset": cardio_preset,
        "category": {"es": GROUP_ES.get(group, group), "en": GROUP_EN.get(group, group)},
        "equipment": {
            "es": EQUIPMENT_ES.get(spec["equipment"], spec["equipment"]),
            "en": EQUIPMENT_EN.get(spec["equipment"], spec["equipment"]),
        },
        "primaryMuscle": {"es": primary_es, "en": primary_en},
        "secondaryMuscles": {"es": secondary_es, "en": secondary_en},
        "names": {"es": spec["name_es"], "en": name_en},
        "descriptions": {"es": spec.get("notes_es", ""), "en": spec.get("notes_en", "")},
        "aliases": {
            "es": spec.get("aliases_es", []),
            "en": spec.get("aliases_en", []),
        },
        "imageUrl": image_url,
        "exercisedbId": exercisedb_id,
        "matchConfidence": match_confidence,
        "matchScore": match_score,
    }


def build_cardio_entry(spec: dict) -> dict:
    group = "Cardio"
    name_en = spec["name_en"]
    ff_id = make_ff_id(group, name_en)
    load_mode = spec["load_mode"]
    secondary = spec.get("secondary", [])
    cardio_preset = spec.get("cardio_preset", "custom")
    logging_type, preset = infer_logging(ff_id, group, name_en, load_mode, cardio_preset)
    exercisedb_id = spec.get("exercisedb_id")
    image_url = gif_url(exercisedb_id)
    if ff_id in LOCAL_IMAGE_OVERRIDES:
        image_url = LOCAL_IMAGE_OVERRIDES[ff_id]
        exercisedb_id = None

    return {
        "id": ff_id,
        "loggingType": logging_type,
        "loadMode": load_mode,
        "perArmWeight": False,
        "unilateral": False,
        "weightOptional": True,
        "cardioPreset": preset,
        "category": {"es": GROUP_ES[group], "en": GROUP_EN[group]},
        "equipment": {
            "es": EQUIPMENT_ES.get(spec["equipment"], spec["equipment"]),
            "en": EQUIPMENT_EN.get(spec["equipment"], spec["equipment"]),
        },
        "primaryMuscle": {"es": "Cardio", "en": "Cardio"},
        "secondaryMuscles": {
            "es": [localize_muscle(m, "es") for m in secondary],
            "en": [localize_muscle(m, "en") for m in secondary],
        },
        "names": {"es": spec["name_es"], "en": name_en},
        "descriptions": {"es": spec.get("notes_es", ""), "en": spec.get("notes_en", "")},
        "aliases": {
            "es": spec.get("aliases_es", []),
            "en": spec.get("aliases_en", []),
        },
        "imageUrl": image_url,
        "exercisedbId": exercisedb_id,
        "matchConfidence": "curated",
        "matchScore": None,
    }


def build_entry(row: tuple, match: dict | None) -> dict:
    group = str(row[0] or "").strip()
    name_en = str(row[1] or "").strip()
    name_es = str(row[2] or "").strip()
    equipment = str(row[3] or "").strip()
    load_mode = str(row[4] or "single_load").strip()
    independent = bool(row[5])
    unilateral = bool(row[6])
    primary = str(row[7] or "").strip()
    secondary = split_secondary(row[8] if len(row) > 8 else None)

    ff_id = make_ff_id(group, name_en)
    logging_type, cardio_preset = infer_logging(ff_id, group, name_en, load_mode)

    per_arm = independent or load_mode == "dual_load"
    weight_optional = load_mode in {"bodyweight", "assisted_bodyweight"}

    exercisedb = (match or {}).get("exercisedb") or {}
    exercisedb_id = exercisedb.get("exerciseId")
    image_url = exercisedb.get("gifUrl")
    match_confidence = (match or {}).get("match_confidence")
    match_score = (match or {}).get("match_score")

    if ff_id in LOCAL_IMAGE_OVERRIDES:
        image_url = LOCAL_IMAGE_OVERRIDES[ff_id]
        exercisedb_id = None
        match_confidence = "curated"
        match_score = None
    elif ff_id in EXERCISEDB_OVERRIDES:
        exercisedb_id = EXERCISEDB_OVERRIDES[ff_id]
        image_url = gif_url(exercisedb_id)
        match_confidence = "curated"
        match_score = None

    primary_es = localize_muscle(primary, "es")
    primary_en = localize_muscle(primary, "en")
    secondary_es = [localize_muscle(m, "es") for m in secondary]
    secondary_en = [localize_muscle(m, "en") for m in secondary]

    return {
        "id": ff_id,
        "loggingType": logging_type,
        "loadMode": load_mode,
        "perArmWeight": per_arm,
        "unilateral": unilateral,
        "weightOptional": weight_optional,
        "cardioPreset": cardio_preset,
        "category": {"es": GROUP_ES.get(group, group), "en": GROUP_EN.get(group, group)},
        "equipment": {
            "es": EQUIPMENT_ES.get(equipment, equipment),
            "en": EQUIPMENT_EN.get(equipment, equipment),
        },
        "primaryMuscle": {"es": primary_es, "en": primary_en},
        "secondaryMuscles": {"es": secondary_es, "en": secondary_en},
        "names": {"es": name_es, "en": name_en},
        "descriptions": {"es": "", "en": ""},
        "aliases": {"es": [], "en": []},
        "imageUrl": image_url,
        "exercisedbId": exercisedb_id,
        "matchConfidence": match_confidence,
        "matchScore": match_score,
    }


def main() -> None:
    matches = load_matches()
    wb = load_workbook(EXCEL, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    LEGACY_OUT.parent.mkdir(parents=True, exist_ok=True)
    if OUT.exists():
        try:
            existing = json.loads(OUT.read_text(encoding="utf-8"))
            if existing.get("version", 1) < 2 and not LEGACY_OUT.exists():
                LEGACY_OUT.write_text(
                    OUT.read_text(encoding="utf-8"),
                    encoding="utf-8",
                )
                print(f"Archived previous catalog to {LEGACY_OUT}")
        except json.JSONDecodeError:
            pass

    exercises = []
    for row in rows[1:]:
        if not row or not row[1]:
            continue
        ff_id = make_ff_id(str(row[0] or ""), str(row[1] or ""))
        if ff_id in EXCLUDE_FF_IDS:
            continue
        exercises.append(build_entry(row, matches.get(ff_id)))

    existing_ids = {e["id"] for e in exercises}
    for spec in CATALOG_SUPPLEMENT:
        entry = build_strength_supplement_entry(spec)
        if entry["id"] not in existing_ids:
            exercises.append(entry)
            existing_ids.add(entry["id"])

    for spec in CARDIO_SUPPLEMENT:
        exercises.append(build_cardio_entry(spec))

    exercises.sort(key=lambda e: e["names"]["en"].lower())

    catalog = {"version": 2, "exerciseCount": len(exercises), "exercises": exercises}
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {len(exercises)} exercises to {OUT}")
    with_image = sum(1 for e in exercises if e.get("imageUrl"))
    print(f"With ExerciseDB GIF: {with_image}")


if __name__ == "__main__":
    main()
